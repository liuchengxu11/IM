from openpyxl import    load_workbook
from Config.Config import Excel_Path
import os


class Get_excel(object):
   def __init__(self):
      self.path=Excel_Path

   def get_excel(self,sheetname):
      shet=None
      print(self.path)
      if os.path.exists(self.path):
         wb= load_workbook(self.path)
         sheetnames=wb.sheetnames
         if sheetnames :
            for setname in sheetnames:
               if setname.lower() == sheetname.lower():
                  shet = wb.active
                  break
            else:
               print("________你要读取的excel里没有{} 这个sheet".format(sheetname))

      else:
         print("------------读取excel文件失败文件不存在请检查路径")

      return  shet
















